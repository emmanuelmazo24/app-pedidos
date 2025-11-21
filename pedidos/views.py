from django.shortcuts import render, redirect, get_object_or_404
from .forms import PedidosForm, DetallePedidosFormSet, DetallePedidosEdFormSet,PreciosIndumentariaForm,MiCambioPasswordForm, EditarPerfilForm
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.db import IntegrityError
from django.contrib.auth.decorators import login_required
from .models import Pedidos, Profile_user,Pedidos_detalle, Pedidos_imagen,PreciosIndumentaria
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.http import HttpResponse, JsonResponse
from django.conf import settings
from .utils import obtener_tipo_usuario
from django.contrib import messages

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

# Create your views here.
def index(request):
    return render(request, 'index.html')

def signup(request):
    if request.method == 'GET':
        return render(request, 'signup.html', {"form": UserCreationForm})
    else:

        if request.POST["password1"] == request.POST["password2"]:
            try:
                user = User.objects.create_user(
                    request.POST["username"], password=request.POST["password1"])
                user.save()
                Profile_user.objects.create(
                    user=user,
                    tipo_usuario = 'CLIENTE'
                )
                login(request, user)
                return redirect('pedidos')
            except IntegrityError:
                return render(request, 'signup.html', {"form": UserCreationForm, "error": "Usuario ya existe."})

        return render(request, 'signup.html', {"form": UserCreationForm, "error": "Verifique password."})  

def signin(request):
    if request.method == 'GET':
        return render(request, 'signin.html', {"form": AuthenticationForm})
    else:
        user = authenticate(
            request, username=request.POST['username'], password=request.POST['password'])
        if user is None:
            return render(request, 'signin.html', {"form": AuthenticationForm, "error": "Usuario o Password incorrecto."})

        #user.tipo_usuario = Profile_user.objects.get(user=user).tipo_usuario if Profile_user.objects.filter(user=user).exists() else 'ADMIN'
        #print(Profile_user.objects.get(user=user).tipo_usuario if Profile_user.objects.filter(user=user).exists() else 'ADMIN')
        #print(user.tipo_usuario)
        login(request, user)
        return redirect('pedidos')

@login_required
def perfil(request):
   return render(request, 'perfil.html')

@login_required
def update_password(request):
    error = None
    if request.method == 'POST':
        form = MiCambioPasswordForm(request.POST)
        if form.is_valid():
            contrasenha_actual = form.cleaned_data['contrasenha_actual']
            nueva_contrasenha = form.cleaned_data['nueva_contrasenha']            
            usuario = request.user

            if not usuario.check_password(contrasenha_actual):
                #error = "La contraseña actual es incorrecta."
                form.add_error('contrasenha_actual', 'La contraseña actual es incorrecta.')
            else:
                usuario.set_password(nueva_contrasenha)
                usuario.save()
                update_session_auth_hash(request, usuario)  # Evita que se cierre la sesión
                messages.success(request, 'Contraseña actualizada con éxito.')
                return redirect('perfil')
        else:
           error = form.non_field_errors()
    else:
        form = MiCambioPasswordForm()
    return render(request, 'cambiar_password.html', {'form': form, 'error' : error})

@login_required
def editar_perfil(request):
    if request.method == 'POST':
        form = EditarPerfilForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado con éxito.')
            return redirect('perfil')
    else:
        form = EditarPerfilForm(instance=request.user)

    return render(request, 'editar_perfil.html', {'form': form})

@login_required
def update_pass(request):
  if request.method == 'POST':
    user = request.user
    if not user.check_password(request.POST["user_pass"]):
        messages.error(request, "La contraseña actual es incorrecta.")
        return render(request, 'perfil.html')
      
    if request.POST["user_pass_nv"] == request.POST["user_pass_nvr"]:
      try:
          user.set_password(request.POST["user_pass_nv"])
          user.save()
          update_session_auth_hash(request, user)  # Evita que se cierre la sesión
          messages.success(request, 'Contraseña actualizada con éxito.')
          #login(request, user)
          return render(request, 'perfil.html')
      except IntegrityError:
        messages.error(request, "Error al validar")
        return render(request, 'perfil.html')
    else:
      messages.error(request, 'Contraseña No son iguales.')
      #login(request, user)
      return render(request, 'perfil.html')

@login_required
def signout(request):
    logout(request)
    return redirect('index')

@login_required
def pedidos(request):
    #profiles = Profile_user.objects.filter(user = request.user)
    tipo_usuario = obtener_tipo_usuario(request.user)
    # if profiles.exists():        
    #     for profile in profiles:
    #         tipo_usuario = profile.tipo_usuario
    #         if profile.tipo_usuario == 'CLIENTE':
    #             pedidos = Pedidos.objects.filter(user=request.user)
    #         else:
    #             pedidos = Pedidos.objects.all()
    # else:
    #     tipo_usuario = 'ADMIN'
    #     pedidos = Pedidos.objects.all()
    if tipo_usuario == 'CLIENTE':
        pedidos = Pedidos.objects.filter(user=request.user)
    else:
        pedidos = Pedidos.objects.all()

    #print(tipo_usuario)
    # if profile.tipo_usuario == 'CLIENTE':
    #     pedidos = Pedidos.objects.filter(user=request.user)
    # else:
    #     pedidos = Pedidos.objects.all()
    return render(request, 'pedidos.html', {'pedidos': pedidos})

@login_required
def crear_pedidos(request):
    if request.method == 'POST':
      try:    
        tipo_usuario = obtener_tipo_usuario(request.user)
        form = PedidosForm(request.POST, request.FILES)
        formset = DetallePedidosFormSet(request.POST)

        if tipo_usuario != 'ADMIN':
            form.fields['senha'].required = False
                            
        if form.is_valid() and formset.is_valid():
            # Contar formularios con datos
            formularios_validos = 0
            for form_data in formset:
                if form_data.cleaned_data and any(form_data.cleaned_data.values()):
                    formularios_validos += 1
            
            if formularios_validos == 0:
                return render(request, 'create_pedido.html', {
                    'form': form, 
                    'formset': formset,
                    'error': 'Debes agregar al menos un detalle del pedido'
                })     
              
            pedidos = form.save(commit=False)
            pedidos.user = request.user
            # upload_result = cloudinary.uploader.upload("",
            #                                 public_id="shoes")
            # upload_result["secure_url"]
            # pedidos.img_prueba = upload_result["secure_url"]
            pedidos.save()
            # Guardar los detalles del pedido
            pedidosDetalle = formset.save(commit=False)
            total_aprobado = 0            
            #un nuevo comentario
            for detalle in pedidosDetalle:
                detalle.pedido = pedidos
                precio = PreciosIndumentaria.objects.get(indumentaria=detalle.indumentaria, calidad=detalle.calidad).precio_unitario
                detalle.precio_aprobado = precio
                detalle.save()
                total_aprobado += precio
            # Guardar la suma en la cabecera del pedido
            pedidos.total = total_aprobado
            pedido.saldo = total_aprobado - pedido.senha
            pedidos.save()
            return redirect('index')  # Redirect to a success page or another view
        else:
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")  

            for form_data in formset:      
                for field, errors in form_data.errors.items():
                    for error in errors:
                        error_messages.append(f"{field}: {error}")                      

            print("Errores:", error_messages)
            return render(request, 'create_pedido.html', {'form': form, 'formset': formset, 'error': 'Por favor, completa el formulario correctamente.','errors_list': error_messages})
      except ValueError as e:
        #print(f"Error al guardar el pedido: {e}")
        form = PedidosForm()
        pedido = Pedidos()
        formset = DetallePedidosFormSet(instance=pedido)
        return render(request, 'create_pedido.html', {'form': form, 'formset': formset, 'error': 'Ocurrió un error al guardar el pedido. Por favor, intente de nuevo.'})
    else:
      form = PedidosForm()
      pedido = Pedidos()
      formset = DetallePedidosFormSet(instance=pedido)      
      return render(request, 'create_pedido.html', {'form': form, 'formset': formset})

@login_required
def pedidos_detalle(request, pedido_id):
  tipo_usuario = obtener_tipo_usuario(request.user)
  if request.method == 'GET':          
    if tipo_usuario == 'CLIENTE':
        pedido = get_object_or_404(Pedidos, pk=pedido_id, user=request.user)
    else:
        pedido = get_object_or_404(Pedidos,pk=pedido_id)

    #pedidos_detalle = Pedidos_detalle.objects.filter(pedido=pedido)
    if pedido.estado == 'PENDIENTE':
        form = PedidosForm(instance=pedido)
        formSet = DetallePedidosEdFormSet(instance=pedido)
        return render(request, 'pedidos_detalle.html', {'form': form,'formSet': formSet, 'pedido': pedido})
    else:
        return redirect('pedidos')
  else:
    try:         
      if tipo_usuario == 'CLIENTE':
          pedido = get_object_or_404(Pedidos, pk=pedido_id, user=request.user)
      else:
          pedido = get_object_or_404(Pedidos,pk=pedido_id)
              
      
      if request.FILES:
          form = PedidosForm(request.POST, request.FILES, instance=pedido)
      else:
          form = PedidosForm(request.POST, instance=pedido)
      total_aprobado = 0
      formset = DetallePedidosEdFormSet(request.POST, instance=pedido)

      if tipo_usuario != 'ADMIN':
            form.fields['senha'].required = False
      # print(form.is_valid())
      # print(formset.is_valid())
      # print(formset.errors)  # Muestra los errores de cada formulario en el formset
      #print(formset.non_form_errors())  # Muestra errores generales del formset
      if form.is_valid() and formset.is_valid():            
          #print('valido')
          form.save()
          # Guardar los detalles del pedido
          pedidosDetalle = formset.save(commit=False)
          #print(pedidosDetalle)
          for detalle in pedidosDetalle:  
              #print('actualiza detalle')              
              detalle.pedido = pedido
              precio = PreciosIndumentaria.objects.get(indumentaria=detalle.indumentaria, calidad=detalle.calidad).precio_unitario
              detalle.precio_aprobado = precio
              detalle.save()                

          pedidosDetalle = Pedidos_detalle.objects.filter(pedido=pedido)
          for detalle in pedidosDetalle:
              total_aprobado += detalle.precio_aprobado
          # Eliminar los detalles marcados para eliminación         
          for detalle in formset.deleted_objects:
              #print(detalle)
              precio = PreciosIndumentaria.objects.get(indumentaria=detalle.indumentaria, calidad=detalle.calidad).precio_unitario
              detalle.delete()
              total_aprobado -= precio
          # Guardar la suma en la cabecera del pedido
          pedido.total = total_aprobado
          pedido.saldo = total_aprobado - pedido.senha
          pedido.save()
      else:
        error_messages = []
        for field, errors in form.errors.items():
            for error in errors:
                error_messages.append(f"{field}: {error}")  

        for form_data in formset:      
            for field, errors in form_data.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")                      

        print("Errores:", error_messages)
        return render(request, 'pedidos_detalle.html', {'form': form, 'formset': formset, 'error': 'Por favor, completa el formulario correctamente.','errors_list': error_messages})
      
      return redirect('pedidos')
    except ValueError:
        return render(request, 'pedidos_detalle.html', {'form': form, 'formSet': formset, 'error': 'Ocurrió un error al actualizar el pedido. Por favor, intente de nuevo.'})

def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html  = template.render(context_dict)
    response = HttpResponse(content_type='application/pdf')
    pisa_status = pisa.CreatePDF(
        html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF: %s' % pisa_status.err)
    return response

@login_required
def pedidos_pdf_view(request, pedido_id):
    #profiles = Profile_user.objects.filter(user = request.user)
    tipo_usuario = obtener_tipo_usuario(request.user)
           
    if tipo_usuario == 'CLIENTE':
        pedidos = get_object_or_404(Pedidos, pk=pedido_id, user=request.user)
    else:
        pedidos = get_object_or_404(Pedidos,pk=pedido_id)

    pedidos_detalle = Pedidos_detalle.objects.filter(pedido=pedidos)
    #pedidos_imagen = Pedidos_imagen.objects.filter(pedido=pedidos)  

    # for img in pedidos_imagen:
    #     img.imagen_path = os.path.join(settings.MEDIA_ROOT, os.path.basename(img.imagen.name))
    pedidos.img_senha_url = request.build_absolute_uri(pedidos.img_senha.url) if pedidos.img_senha else ''
    pedidos.img_jugadores_url = request.build_absolute_uri(pedidos.img_jugadores.url) if pedidos.img_jugadores else ''
    pedidos.img_arquero_url = request.build_absolute_uri(pedidos.img_arquero.url) if pedidos.img_arquero else ''
    pedidos.img_auspicio1_url = request.build_absolute_uri(pedidos.img_auspicio1.url) if pedidos.img_auspicio1 else ''
    pedidos.img_auspicio2_url = request.build_absolute_uri(pedidos.img_auspicio2.url) if pedidos.img_auspicio2 else ''
    pedidos.img_auspicio3_url = request.build_absolute_uri(pedidos.img_auspicio3.url) if pedidos.img_auspicio3 else ''
    pedidos.img_auspicio4_url = request.build_absolute_uri(pedidos.img_auspicio4.url) if pedidos.img_auspicio4 else ''
    pedidos.img_auspicio5_url = request.build_absolute_uri(pedidos.img_auspicio5.url) if pedidos.img_auspicio5 else ''

    # for img in pedidos_imagen:
    #     img.imagen_url = request.build_absolute_uri(img.imagen.url)  

    total_montos = sum([d.precio_aprobado * d.cantidad for d in pedidos_detalle])
    
    context = {
        "pedidos": pedidos,
        "pedidos_detalle": pedidos_detalle,        
        "total_montos": total_montos,
    }
    #return render_to_pdf('pedido_pdf_template.html', context)
    return render(request, 'pedido_pdf_template.html', context)

@login_required
def obtener_precio(request):
    indumentaria = request.GET.get('indumentaria')
    calidad = request.GET.get('calidad')
    talle = request.GET.get('talle')
    lis_indumentaria_remera = ['EQUIPO COMPLETO','CAMISETA SOLA']
    lis_indumentaria_short = ['SHORT SOLA']
    try:
        precio = PreciosIndumentaria.objects.get(indumentaria=indumentaria, calidad=calidad).precio_unitario
        if talle in ['XL','XXL','3XL']:
            if indumentaria in lis_indumentaria_remera:
                precio = precio + 10000
            if indumentaria in lis_indumentaria_short:
                precio = precio + 5000
    except PreciosIndumentaria.DoesNotExist:
        precio = 0
    return JsonResponse({'precio': precio})

@login_required
def pedidos_aprobar(request, pedido_id):
    pedido = get_object_or_404(Pedidos, pk=pedido_id)
    if request.method == 'POST':
        pedido.estado = 'APROBADO' if pedido.estado == 'PENDIENTE' else 'PENDIENTE'
        pedido.save()
        return redirect('pedidos')

@login_required
def precio_indumentaria(request,precio_id=None):
  error = None
  if precio_id:
    precio = get_object_or_404(PreciosIndumentaria,pk=precio_id)
  else:
    precio = None

  if request.method == 'POST':    
    form = PreciosIndumentariaForm(request.POST, instance=precio)
    if form.is_valid():
        precio_form = form.save(commit=False)
        precio_form.user = request.user
        precio_form.save()
        return redirect('precio_indumentaria') # Redirige al listado
    else:
        error = "Ocurrió un error al procesar el formulario."
  else:
    form = PreciosIndumentariaForm(instance=precio) 

  precios = PreciosIndumentaria.objects.all() 
  context = {
      'form': form,
      'precios': precios,
      'error': error,
    }
  return render(request, 'precios_indumentaria.html', context)

@login_required
def del_precio(request, precio_id):
  tipo_usuario = obtener_tipo_usuario(request.user)
  if tipo_usuario != 'CLIENTE':
    precios = get_object_or_404(PreciosIndumentaria, pk=precio_id)
    if request.method == 'POST':
        precios.delete()
        return redirect('precio_indumentaria')

def crear_superusuario(request):
    if User.objects.filter(username='admin').exists():
        return HttpResponse("El superusuario ya existe.")

    User.objects.create_superuser(
        username='admin',
        email='admin@example.com',
        password='admin123'  # Usa una contraseña más segura en producción
    )
    return HttpResponse("Superusuario creado.")

@login_required
def exportar_detalle_excel_openpyxl(request,pedido_id):
    """Exporta datos a Excel usando openpyxl"""
    
    # Obtener datos
    tipo_usuario = obtener_tipo_usuario(request.user)
           
    if tipo_usuario == 'CLIENTE':
        pedidos = get_object_or_404(Pedidos, pk=pedido_id, user=request.user)
    else:
        pedidos = get_object_or_404(Pedidos,pk=pedido_id)

    pedidos_detalle = Pedidos_detalle.objects.filter(pedido=pedidos)    
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Detalle del Pedido Nro. {pedido_id}"
    
    # Definir estilos
    encabezado_font = Font(bold=True, color="FFFFFF")
    encabezado_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    encabezado_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Agregar encabezados
    encabezados = ['NOMBRE', 'TALLE', 'NRO DORSAL', 'MOLDE', 'TIPO CUELLO', 'COLOR CUELLO', 'TIPO JUGADOR', 'OBS', 'INDUMENTARIA', 'CALIDAD', 'CANTIDAD']
    ws.append(encabezados)
    
    # Aplicar estilos a encabezados
    for cell in ws[1]:
        cell.font = encabezado_font
        cell.fill = encabezado_fill
        cell.alignment = encabezado_alignment
        cell.border = border
    
    # Agregar datos
    for detalle in pedidos_detalle:
        ws.append([
            detalle.nombre,
            detalle.talle,
            detalle.dorsal, #fecha.strftime('%d/%m/%Y')
            detalle.molde,
            detalle.cuello_tipo,
            detalle.cuello_color,
            detalle.tipo_jugador,
            detalle.observacion,
            detalle.indumentaria,
            detalle.calidad,
            detalle.cantidad
            #f"${detalle.monto_total:.2f}",
            #detalle.estado
        ])
    
    # Aplicar estilos a datos
    for row in ws.iter_rows(min_row=2, max_row=len(pedidos_detalle)+1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 12

    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="detalle_pedido_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response